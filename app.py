# -*- coding: utf-8 -*-
"""
Веб-дашборд статистики ключевых слов Yandex Wordstat.

Бэкенд проксирует запросы к Yandex Search API, строго разбивает данные по
субъектам РФ (федеральные округа исключаются) и формирует выгрузку в Excel.
"""

import io
import os
import sys

from flask import Flask, jsonify, request, send_file

from months import last_n_months


def _load_dotenv():
    """Минимальная загрузка .env в окружение (без внешних зависимостей)."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


_load_dotenv()

import dataset
import region_weights
from regions import build_subject_index
from yandex_client import WordstatClient, YandexError

app = Flask(__name__, static_folder="static", static_url_path="")

# Индекс субъектов кэшируется на время жизни процесса.
_subject_index_cache = {"value": None}

# Если API ответил ошибкой — не мучаем его ретраями на каждый запрос,
# а на время уходим в локальные данные (предохранитель).
_api_down_until = {"t": 0.0}
_API_COOLDOWN = 120  # сек


def get_subject_index(client):
    if _subject_index_cache["value"] is None:
        try:
            tree = client.get_regions_tree()
        except YandexError:
            tree = None
        _subject_index_cache["value"] = build_subject_index(tree)
    return _subject_index_cache["value"]


@app.route("/")
def index():
    # send_static_file резолвит путь от root_path приложения (каталог app.py),
    # поэтому работает независимо от текущего рабочего каталога процесса.
    return app.send_static_file("index.html")


@app.route("/api/status")
def status():
    client = WordstatClient()
    return jsonify({"demo": client.demo_mode})


@app.route("/api/products")
def products():
    """Фразы с реальными данными по тоталу РФ (из ручной выгрузки CSV)."""
    return jsonify({"phrases": dataset.known_phrases()})


@app.route("/api/search", methods=["POST"])
def search():
    payload = request.get_json(force=True, silent=True) or {}
    phrase = (payload.get("phrase") or "").strip()
    if not phrase:
        return jsonify({"error": "Не задана ключевая фраза"}), 400

    devices = payload.get("devices") or None
    import time as _time
    cooled = _time.time() < _api_down_until["t"]
    client = WordstatClient(force_demo=cooled)
    index = get_subject_index(client)

    try:
        key, base, otp_name, hist, year, _my, _oy = \
            _phrase_history(phrase, devices, client, index)
    except YandexError:
        # API недоступен (права/квота/сеть) — не роняем сайт, а прозрачно
        # откатываемся на локальные данные (CSV для известных продуктов).
        _api_down_until["t"] = _time.time() + _API_COOLDOWN
        client = WordstatClient(force_demo=True)
        key, base, otp_name, hist, year, _my, _oy = \
            _phrase_history(phrase, devices, client, index)

    # Регионы (числа/доли/лидеры) считаются на клиенте под выбранный период.
    # Рабочий режим: веса = реальное распределение фразы по субъектам из API;
    # демо (или сбой API): модель по населению/спросу.
    subjects = None
    if not client.demo_mode:
        subjects = _live_subjects(client, base, devices, index)
    if subjects is None:
        names = list(index["subjects"].values())
        ws = region_weights.weights_for(names)
        subjects = [{"id": sid, "name": name, "weight": w}
                    for (sid, name), w in zip(index["subjects"].items(), ws)]

    # Конкурентный анализ: бренды в категории (масштаб по устройству как у ОТП).
    of = _device_factors(devices)[1]
    competitors = []
    for brand in dataset.competitor_series(key):
        competitors.append({
            "brand": brand["brand"], "isOtp": brand["isOtp"],
            "series": [int(round(v * of)) for v in brand["series"]],
        })

    # Источник: реальная выгрузка (CSV) для известных продуктов, иначе — демо.
    source = "csv" if (client.demo_mode and dataset.lookup(phrase)) else (
        "api" if not client.demo_mode else "synthetic")

    # Реальный мобильный срез (для графика «Mobile-спрос»): только в рабочем
    # режиме и только когда выбраны «Все устройства» (иначе знаменатель не ALL).
    mobile = None
    if not client.demo_mode and _api_devices(devices) is None:
        try:
            mobile = {
                "market": client.monthly_history(base, hist["keys"],
                                                 devices=["DEVICE_PHONE"]),
                "otp": client.monthly_history(otp_name, hist["keys"],
                                              devices=["DEVICE_PHONE"]),
            }
        except YandexError:
            mobile = None

    return jsonify({
        "marketPhrase": base,
        "otpPhrase": otp_name,
        "demo": client.demo_mode,
        "source": source,
        "queryPhrase": phrase,
        "knownPhrases": dataset.known_phrases(),
        "year": year,
        "subjects": subjects,
        "competitors": competitors,
        # Реальный брендовый спрос «ОТП Банк» (для SoS-аналитики банков):
        # демо — из выгрузки CSV, рабочий режим — из API за те же 48 месяцев.
        "brandOtp": _brand_otp_series(client, hist),
        # Мобильный срез рынка и ОТП (только рабочий режим, «Все устройства»).
        "mobile": mobile,
        "dynamics": hist,            # полная месячная история {keys,labels,market,otp}
        "excluded": {
            "federalDistricts": len(index["federal_districts"]),
            "note": "Федеральные округа и зарубежные регионы исключены.",
        },
    })


@app.route("/api/export", methods=["POST"])
def export_excel():
    """Расширенная выгрузка: Сводка (KPI + выводы), Динамика, Регионы,
    Конкуренты, Прогноз, Сезонность — всё, что видно на дашборде."""
    payload = request.get_json(force=True, silent=True) or {}
    phrase = (payload.get("phrase") or "запрос").strip()
    buffer = _build_workbook(payload)
    filename = "wordstat_%s.xlsx" % _safe_name(phrase)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet",
    )


# ----------------------------------------------------------- helpers ---
# Доли устройств (демо): «Все устройства» = полный тотал (совпадает с CSV),
# остальные — детерминированная доля рынка/ОТП, чтобы селектор реально менял
# цифры. ОТП слегка иначе распределён по устройствам, чем рынок.
_DEVICE_FACTORS = {
    "DEVICE_ALL": (1.0, 1.0),
    "DEVICE_DESKTOP": (0.42, 0.36),
    "DEVICE_PHONE": (0.50, 0.57),
    "DEVICE_TABLET": (0.08, 0.07),
}


def _device_factors(devices):
    if not devices:
        return (1.0, 1.0)
    return _DEVICE_FACTORS.get(str(devices[0]).upper(), (1.0, 1.0))


def _api_devices(devices):
    """Devices для API: None = все устройства (DEVICE_ALL не передаём)."""
    if not devices:
        return None
    d = [str(x).upper() for x in devices if x]
    if not d or d == ["DEVICE_ALL"]:
        return None
    return d


def _phrase_history(phrase, devices, client, index):
    """Единый источник данных по фразе.

    Возвращает (key, base, otpName, hist, year, market_year, otp_year), где
    hist — полная месячная история {keys,labels,market,otp}, а *_year — суммы
    за последний полный календарный год.

    Рабочий режим (есть ключ API): обе истории берутся из Wordstat API за
    48 месяцев, срез по устройствам делает сам API. Демо: известные продукты —
    из CSV, остальные — синтетика; устройства масштабируются множителями.
    """
    ds_key = dataset.lookup(phrase) if client.demo_mode else None
    if ds_key:
        mf, of = _device_factors(devices)
        product = dataset.product(ds_key)
        h = dataset.history(ds_key)
        keys, labels = h["keys"], h["labels"]
        market = [int(round(v * mf)) for v in h["market"]]
        otp = [int(round(v * of)) for v in h["otp"]]
        key, base, otp_name = ds_key, product["base"], product["otpName"]
    else:
        otp_phrase = phrase + " ОТП"
        win = last_n_months(48)
        keys, labels = win["keys"], win["labels"]
        if client.demo_mode:
            mf, of = _device_factors(devices)
            market = [int(round(v * mf)) for v in client.monthly_history(phrase, keys)]
            otp = [int(round(v * of)) for v in client.monthly_history(otp_phrase, keys)]
        else:
            dv = _api_devices(devices)
            market = client.monthly_history(phrase, keys, devices=dv)
            otp = client.monthly_history(otp_phrase, keys, devices=dv)
        key, base, otp_name = phrase, phrase, otp_phrase

    hist = {"keys": keys, "labels": labels, "market": market, "otp": otp}
    year, my, oy = dataset.year_totals_from_series(keys, market, otp)
    return key, base, otp_name, hist, year, my, oy


def _brand_otp_series(client, hist):
    """Брендовый спрос «ОТП Банк» по месяцам hist["keys"] (для SoS-блока)."""
    if client.demo_mode:
        return dataset.otp_brand_series()
    try:
        return client.monthly_history("ОТП Банк", hist["keys"])
    except YandexError:
        return []


def _live_subjects(client, phrase, devices, index):
    """Реальные веса субъектов РФ: распределение фразы по регионам из API.

    Возвращает None при ошибке API или пустом распределении — тогда
    вызывающий код откатывается на модельные веса.
    """
    try:
        resp = client.get_regions_distribution(phrase, devices=_api_devices(devices))
    except YandexError:
        return None
    counts = {}
    for r in resp.get("results") or []:
        rid = str(r.get("region"))
        if rid in index["subjects"]:
            counts[rid] = counts.get(rid, 0) + _safe_int(r.get("count"))
    total = sum(counts.values())
    if not total:
        return None
    return [{"id": sid, "name": name, "weight": counts.get(sid, 0) / total}
            for sid, name in index["subjects"].items()]


def _merge_regions(market_regions, otp_regions):
    """Сводит распределения рынка и ОТП в строки по субъектам РФ.

    Для каждого региона: market/otp (абсолют), marketShare/otpShare (доля
    внутри своего набора) и penetration — доля ОТП от рынка в этом регионе.
    """
    otp_by_id = {r["id"]: r for r in otp_regions}
    total_m = sum(r["count"] for r in market_regions) or 1
    total_o = sum(r["count"] for r in otp_regions) or 1
    rows = []
    for m in market_regions:
        mc = m["count"]
        oc = int(otp_by_id.get(m["id"], {}).get("count", 0))
        rows.append({
            "id": m["id"],
            "name": m["name"],
            "market": mc,
            "otp": oc,
            "marketShare": round(mc / total_m * 100, 2),
            "otpShare": round(oc / total_o * 100, 2),
            "penetration": round(oc / mc * 100, 3) if mc else 0.0,
            "affinityIndex": round(m.get("affinityIndex", 0)),
        })
    rows.sort(key=lambda r: r["market"], reverse=True)
    return rows


def _build_workbook(p):
    """Книга из 6 листов по payload дашборда: Сводка, Динамика, Регионы,
    Конкуренты, Прогноз, Сезонность. Пустые секции пропускаются."""
    import datetime as _dt

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    GREEN, DARK = "52AE30", "20262E"
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor=GREEN)
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color=DARK)
    label_font = Font(bold=True, color=DARK)
    NUM, PCT = "#,##0", "0.000%"

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def fill_table(ws, headers, rows, widths, num_cols=(), pct_cols=()):
        ws.append(headers)
        style_header(ws, 1, len(headers))
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w
        for row in ws.iter_rows(min_row=2, max_row=1 + len(rows)):
            for cell in row:
                if cell.column in num_cols:
                    cell.number_format = NUM
                elif cell.column in pct_cols:
                    cell.number_format = PCT
        ws.freeze_panes = "A2"

    kpi = p.get("kpi") or {}
    dyn = p.get("dynamics") or {}
    labels = dyn.get("labels") or []
    market = dyn.get("market") or []
    otp = dyn.get("otp") or []

    # --- Лист 1: Сводка (KPI + ключевые выводы) ---
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Аналитика Wordstat · рынок и бренд ОТП"
    ws["A1"].font = title_font
    ws["A2"] = "Фраза (рынок): %s" % (p.get("phrase") or "")
    ws["A3"] = "Фраза (бренд): %s" % (p.get("otpPhrase") or "")
    ws["A4"] = "Сформировано: %s" % _dt.datetime.now().strftime("%d.%m.%Y %H:%M")
    row = 6
    pairs = [
        ("Диапазон анализа", kpi.get("range")),
        ("Гранулярность", kpi.get("granularity")),
        ("Последний период", kpi.get("periodLabel")),
        ("Количество запросов · рынок", kpi.get("market")),
        ("Количество запросов · ОТП", kpi.get("otp")),
        ("Доля ОТП от рыночных запросов", (kpi.get("share") or 0) / 100.0),
        ("Динамика · рынок", kpi.get("growthMarket")),
        ("Динамика · ОТП", kpi.get("growthOtp")),
        ("SoS · доля голоса ОТП", kpi.get("sos")),
        ("SoS · изменение за период", kpi.get("sosDelta")),
        ("Лидер роста SoS", kpi.get("sosWinner")),
        ("Брендовые запросы ОТП", kpi.get("brandOtp")),
        ("Интенты · доля лидов", kpi.get("intentLeads")),
        ("Интенты · сервисные", kpi.get("intentService")),
        ("Интенты · токсичные", kpi.get("intentToxic")),
        ("Индекс сезонности запросов ОТП", kpi.get("seasonIdx")),
    ]
    for name, value in pairs:
        ws.cell(row=row, column=1, value=name).font = label_font
        cell = ws.cell(row=row, column=2, value=value)
        if name.startswith("Количество"):
            cell.number_format = NUM
        if name.startswith("Доля"):
            cell.number_format = PCT
        row += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40

    # --- Лист 2: Динамика (рынок, ОТП, доля) ---
    if labels:
        ws2 = wb.create_sheet("Динамика")
        rows = []
        for i, lab in enumerate(labels):
            m = market[i] if i < len(market) else None
            o = otp[i] if i < len(otp) else None
            share = (o / m) if (m and o is not None) else None
            rows.append([lab, m, o, share])
        fill_table(ws2, ["Период", "Рынок, запросов", "ОТП, запросов", "Доля ОТП"],
                   rows, [16, 16, 15, 12], num_cols=(2, 3), pct_cols=(4,))
        line = LineChart()
        line.title = "Динамика запросов: рынок и ОТП"
        line.height, line.width = 10, 26
        data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=1 + len(rows))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(rows))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ws2.add_chart(line, "F2")

    # --- Лист 3: Регионы (сортировка — по Value Pool, как на дашборде) ---
    regions = p.get("regions") or []
    if regions:
        ws3 = wb.create_sheet("Регионы")
        rows = [[r.get("name"), r.get("market"), r.get("otp"),
                 (r.get("penetration") or 0) / 100.0,
                 r.get("pool"),
                 (r.get("marketShare") or 0) / 100.0,
                 r.get("affinityIndex")] for r in regions]
        fill_table(ws3, ["Регион (субъект РФ)", "Рынок, запросов", "ОТП, запросов",
                         "Доля ОТП от рынка", "Value Pool, запросов",
                         "Доля региона в рынке", "Индекс соответствия"],
                   rows, [38, 16, 15, 17, 19, 19, 18],
                   num_cols=(2, 3, 5), pct_cols=(4, 6))
        n = min(len(regions), 15)
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Топ-15 регионов: рынок и ОТП"
        chart.height, chart.width = 12, 24
        data = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=1 + n)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=1 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws3.add_chart(chart, "H2")

    # --- Лист 4: Конкуренты ---
    comp = p.get("competitors") or []
    if comp and labels:
        ws4 = wb.create_sheet("Конкуренты")
        headers = ["Период"] + [c.get("brand") for c in comp]
        rows = []
        for i, lab in enumerate(labels):
            rows.append([lab] + [(c.get("series") or [None])[i]
                                 if i < len(c.get("series") or []) else None
                                 for c in comp])
        fill_table(ws4, headers, rows, [16] + [14] * len(comp),
                   num_cols=tuple(range(2, len(comp) + 2)))
        line = LineChart()
        line.title = "Брендовые запросы: ОТП и конкуренты"
        line.height, line.width = 10, 26
        data = Reference(ws4, min_col=2, max_col=1 + len(comp),
                         min_row=1, max_row=1 + len(rows))
        cats = Reference(ws4, min_col=1, min_row=2, max_row=1 + len(rows))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ws4.add_chart(line, get_column_letter(len(comp) + 3) + "2")

    # --- Лист 5: Прогноз доли ОТП ---
    fc = p.get("forecast") or {}
    if fc.get("labels"):
        ws5 = wb.create_sheet("Прогноз")
        rows = []
        for i, lab in enumerate(fc["labels"]):
            def val(key):
                arr = fc.get(key) or []
                v = arr[i] if i < len(arr) else None
                return (v / 100.0) if v is not None else None
            rows.append([lab, val("actual"), val("base"), val("opt"), val("pess")])
        fill_table(ws5, ["Период", "Факт", "Базовый", "Оптимистичный", "Пессимистичный"],
                   rows, [16, 12, 12, 15, 17], pct_cols=(2, 3, 4, 5))
        line = LineChart()
        line.title = "Прогноз доли ОТП (3 сценария)"
        line.height, line.width = 10, 26
        data = Reference(ws5, min_col=2, max_col=5, min_row=1, max_row=1 + len(rows))
        cats = Reference(ws5, min_col=1, min_row=2, max_row=1 + len(rows))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ws5.add_chart(line, "H2")

    # --- Лист 6: Интенты (структура брендовых запросов ОТП) ---
    intents = p.get("intents") or {}
    if intents.get("labels"):
        ws_i = wb.create_sheet("Интенты")
        rows = []
        for i, lab in enumerate(intents["labels"]):
            r = (intents.get("rows") or [[]])[i] if i < len(intents.get("rows") or []) else []
            vals = [r[j] if j < len(r) else None for j in range(4)]
            tot = sum(v for v in vals if v) or 1
            rows.append([lab] + vals + [v / tot if v is not None else None for v in vals])
        fill_table(ws_i, ["Месяц", "Продуктовый (лиды)", "Сервисный", "Токсичный", "Прочее",
                          "Лиды, %", "Сервис, %", "Токсичные, %", "Прочее, %"],
                   rows, [14, 17, 13, 13, 12, 11, 12, 13, 11],
                   num_cols=(2, 3, 4, 5), pct_cols=(6, 7, 8, 9))

    # --- Лист 7: Сезонность (матрицы год × месяц) ---
    sea = p.get("seasonality") or {}
    if sea.get("years"):
        ws6 = wb.create_sheet("Сезонность")
        months = sea.get("months") or []
        for block, key in (("Рынок: запросы по месяцам", "market"),
                           ("ОТП: запросы по месяцам", "otp")):
            ws6.append([block])
            ws6.cell(row=ws6.max_row, column=1).font = title_font
            ws6.append(["Год"] + months)
            style_header(ws6, ws6.max_row, len(months) + 1)
            for yi, year in enumerate(sea["years"]):
                vals = (sea.get(key) or [[]])[yi] if yi < len(sea.get(key) or []) else []
                ws6.append([year] + [vals[i] if i < len(vals) else None
                                     for i in range(len(months))])
                for cell in ws6[ws6.max_row][1:]:
                    cell.number_format = NUM
            ws6.append([])
        ws6.column_dimensions["A"].width = 10
        for i in range(len(months)):
            ws6.column_dimensions[get_column_letter(i + 2)].width = 11

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _safe_int(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _safe_name(phrase):
    keep = [c if c.isalnum() else "_" for c in phrase.lower()]
    return ("".join(keep)[:40]) or "export"


def _resolve_port():
    """Порт: --port N → переменная PORT → 5000 по умолчанию.

    На macOS порт 5000 по умолчанию занимает AirPlay (IPv6 ::1), поэтому
    возможность переопределить порт особенно полезна локально.
    """
    argv = sys.argv
    if "--port" in argv:
        try:
            return int(argv[argv.index("--port") + 1])
        except (IndexError, ValueError):
            pass
    return int(os.environ.get("PORT", "5000"))


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=_resolve_port(), debug=False)
